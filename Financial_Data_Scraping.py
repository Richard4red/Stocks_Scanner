#!/usr/bin/env python
# coding: utf-8

# In[27]:


import openpyxl
import requests
from bs4 import BeautifulSoup
import pandas as pd
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import os
import numpy as np
import threading
import queue


# In[19]:


def rate(pv, fv, n):
    r = (fv / pv) ** (1/n) - 1
    return r


# In[1]:


def company_evaluation(ticker):
    urls = {}
    urls['income sheet'] = f"https://stockanalysis.com/stocks/{ticker}/financials/"
    urls['balance sheet'] = f"https://stockanalysis.com/stocks/{ticker}/financials/balance-sheet/"
    urls['cash flow'] = f"https://stockanalysis.com/stocks/{ticker}/financials/cash-flow-statement/"
    urls['PE'] = f"https://stockanalysis.com/stocks/{ticker.lower()}/financials/ratios/?p=quarterly"
    
    dataframe_dict = {}
    
    xlwriter = pd.ExcelWriter(f'financial statements ({ticker}).xlsx', engine='xlsxwriter')
    
    with open('valid_proxies.txt',"r") as f: 
        proxies = f.read().split('\n')
    
    #inorder to not get banned from  the website Im scraping from I will use Proxies
    #flag is to indicate that everyting went well and we can continue with the for loop
    
    proxy_index = 0 
    flag = True
    
    try:
        for statement ,url in urls.items():
            flag = True
            if statement == 'PE':
                while flag:    
                    try:
                        response = requests.get(url, proxies = {'http':proxies[proxy_index],'https':proxies[proxy_index]})
                        print(f"Using the Proxy: {proxies[proxy_index]}")
                        soup = BeautifulSoup(response.content, 'html.parser')
                        df_pe = pd.read_html(str(soup), attrs={'data-test': 'financials'}, index_col = 0)[0]
                        df_pe = df_pe.iloc[:, :-1]
                        df_pe = df_pe.loc['PE Ratio']
                        dataframe_dict[statement] = df_pe
                        flag = False
                        
                    except:
                        print(f"Faild Proxy: {proxies[proxy_index]}")
                        
                    finally:
                        proxy_index += 1
                        proxy_index %= len(proxies)
                        
            else:     
                while flag:
                    try:
                        chrome_options = webdriver.ChromeOptions()
                        chrome_options.add_argument('--proxy-server=%s' % proxies[proxy_index])

                        print(f"Using the Proxy: {proxies[proxy_index]}")

                        driver = webdriver.Chrome()
                        driver.get(url)

                        button = driver.find_element(By.XPATH, "//button[@class='controls-btn']")
                        button.click()

                        raw_option = driver.find_element(By.XPATH, "//div[contains(@class, 'absolute') and contains(@class, 'z-50')]//button[contains(text(), 'Raw')]")
                        raw_option.click()

                        wait = WebDriverWait(driver, 10)
                        table = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.CSS_SELECTOR, '[data-test="financials"]')))
                        temp = pd.read_html(table.get_attribute('outerHTML'), index_col = 0)[0]
                        temp = temp.iloc[:, :-1]
                        temp = temp.iloc[:, ::-1]
                        dataframe_dict[statement] = temp
                        driver.quit()
                        flag = False
                        
                    except:
                        print(f"Faild Proxy: {proxies[proxy_index]}")
                        
                    finally:
                        proxy_index += 1
                        proxy_index %= len(proxies)
            
    
        dataframe_dict['income sheet'].loc["Shareholders' Equity"] = dataframe_dict['balance sheet'].loc["Shareholders' Equity"]
        #temp = dataframe_dict['income sheet'].loc['EPS (Diluted)']
        dataframe_dict['income sheet'].loc['NetCashProvidedByOperatingActivities'] = dataframe_dict['cash flow'].loc['Operating Cash Flow']
        Report = dataframe_dict['income sheet'].loc[["Shareholders' Equity",'NetCashProvidedByOperatingActivities','Revenue','Net Income','Shares Outstanding (Diluted)']]

        #Converting all Columns to number types
        Report.loc["Shareholders' Equity"] = pd.to_numeric(Report.loc["Shareholders' Equity"])
        Report.loc["NetCashProvidedByOperatingActivities"] = pd.to_numeric(Report.loc["NetCashProvidedByOperatingActivities"])
        Report.loc['Revenue'] = pd.to_numeric(Report.loc['Revenue'])
        Report.loc['Net Income'] = pd.to_numeric(Report.loc['Net Income'])
        Report.loc["Shares Outstanding (Diluted)"] = pd.to_numeric(Report.loc["Shares Outstanding (Diluted)"])
        #we didnt add EPS Diluted because of the order we want it to be
        dataframe_dict['income sheet'].loc['EPS (Diluted)'] = pd.to_numeric(dataframe_dict['income sheet'].loc['EPS (Diluted)'])


        Report.loc['Book_Value_Per_Share'] = Report.loc["Shareholders' Equity"] / Report.loc["Shares Outstanding (Diluted)"]
        Report.loc['Book_Value_Per_Share_GROWTH'] = Report.loc['Book_Value_Per_Share'].pct_change()   

        Report.loc['OCPS'] = Report.loc["NetCashProvidedByOperatingActivities"] / Report.loc["Shares Outstanding (Diluted)"]
        Report.loc['OCPS_GROWTH'] = Report.loc['OCPS'].pct_change()

        Report.loc['Sales_Per_Share'] = Report.loc['Revenue'] / Report.loc["Shares Outstanding (Diluted)"]
        Report.loc['Sales_Per_Share_GROWTH'] = Report.loc['Sales_Per_Share'].pct_change() 

        Report.loc['EPS'] = dataframe_dict['income sheet'].loc['EPS (Diluted)']
        Report.loc['EPS_GROWTH'] = Report.loc['EPS'].pct_change()

        Report.loc['ROE'] = Report.loc['Net Income'] / Report.loc["Shareholders' Equity"]

        #Befeore I will formate ROE I will to store it in raw Numbers neccesarry for the next step
        #Prepring For the Calculation Step

        average_list_ROE = [Report.loc['ROE','2013':'2022'].mean(), Report.loc['ROE','2015':'2022'].mean(), Report.loc['ROE','2017':'2022'].mean(), Report.loc['ROE','2019':'2022'].mean(), Report.loc['ROE','2021':'2022'].mean()]

        # Formating Numbers to Precentages
        Report.loc['Book_Value_Per_Share_GROWTH'] = Report.loc['Book_Value_Per_Share_GROWTH'].apply(lambda x: '{:.2%}'.format(x))
        Report.loc['OCPS_GROWTH'] = Report.loc['OCPS_GROWTH'].apply(lambda x: '{:.2%}'.format(x))
        Report.loc['Sales_Per_Share_GROWTH'] = Report.loc['Sales_Per_Share_GROWTH'].apply(lambda x: '{:.2%}'.format(x))
        Report.loc['EPS_GROWTH'] = Report.loc['EPS_GROWTH'].apply(lambda x: '{:.2%}'.format(x))
        Report.loc['ROE'] = Report.loc['ROE'].apply(lambda x: '{:.2%}'.format(x))

        driver.quit()
        Report.to_excel(xlwriter, sheet_name="Company Evaluation", index=True)

        print(f"Finished Getting Information about {ticker}")
        
        #Continuing to the Next Step the Calculations
        company_calculation(Report, dataframe_dict['PE'], average_list_ROE, ticker, xlwriter)

    except Exception as e:
        print(f"An error occurred: with {ticker} the website don't have the relevant data")
        Report.to_excel(xlwriter, sheet_name="Company Evaluation", index=True)
        xlwriter.close()
        
        


# In[4]:


def company_calculation(df, df_pe, average_list_ROE ,ticker, xlwriter):
    
    years = [9, 7, 5, 3, 1]
    rows = ['Book_Value_Per_Share', 'OCPS', 'Sales_Per_Share', 'EPS']
    rate_dict = {}
    evaluation = pd.DataFrame(columns = years)
    
    try:
        for row in rows:
            pv_list = [df.loc[row,'2013'], df.loc[row,'2015'], df.loc[row,'2017'], df.loc[row,'2019'], df.loc[row,'2021']]
            fv_list = [df.loc[row,'2022']] * len(years)
            #applying rate function element wise with list
            rate_dict[row] = [rate(pv_list[i], fv_list[i], years[i]) for i in range(len(pv_list))]
            #preparing evaluation dataframe to check it in attention_stocks function
            evaluation.loc[row] = np.array(rate_dict[row]) * 100
            #Formating all the Numbers to Precentages
            rate_dict[row] = ["{:.2%}".format(x) for x in rate_dict[row]]

        evaluation.loc['ROE'] = np.array(average_list_ROE) * 100
        rows.append('ROE')

        #we want pay attention to companies with stats that are greater than 10% 
        #(the precentages are already look like 0.1 = 10% we returning them to 10% = 10 by * 100)
        #then we counting each row to see if atleast 4 numbers are greater than 10

        attention_stocks(evaluation, rows, ticker)

        formated_mean_ROE = ['{:.2%}'.format(x) for x in average_list_ROE]

        #Headers for the Excel sheet
        headers = ['BVPS GROWTH', 'OCPS GROWTH', 'Sales per share GROWTH', 'EPS GROWTH', 'ROE']
        rate_dict['ROE'] = formated_mean_ROE
        CalculatedReport = pd.DataFrame(rate_dict.values() , columns = years, index = headers)

        CalculatedReport.to_excel(xlwriter, sheet_name="Rate Evaluation", index=True)

        #And for the last sheet we exporting PE ratio to Excel
        df_pe.to_excel(xlwriter, sheet_name="PE Ratio", index=True)

        print(f"Finished Calculations with {ticker}")
        
    except:
        print("We can't calculate rate values because this company has negative values OR not enough years to work with")
        print("Move on to the Next Company this company not good for long term investment right now")
        
    finally:
        xlwriter.close()


# In[5]:


def attention_stocks(evaluation, rows, ticker):
    evaluationFlags = []
    evaluation = evaluation > 10
    
    for row in rows:
        if evaluation.loc[row].sum() >= 4:
            evaluationFlags.append(True)
        else:
            evaluationFlags.append(False)
            
    print("Checking if row has atleast 4 number greater than 10")
    print(evaluationFlags)
    print("-------------------------------------------------------------")
    if all(evaluationFlags):
        
        print(f"Pay attention to {ticker}")
        
        if(not os.path.exists('Worthy_Stocks.txt')):
            with open('Worthy_Stocks.txt', 'w') as file:
                file.write(ticker+'\n')
        else:
            with open('Worthy_Stocks.txt', 'a+') as file:
                file.write(ticker+'\n')


# In[24]:


def company_evaluation_Modification(ticker):
    #loading the file and the desired sheet
    wb = openpyxl.load_workbook(f'financial statements ({ticker}).xlsx')
    sheet = wb["Company Evaluation"]
    
    #Looking for Growth rowes in the sheet
    growth_rows = [row_idx for row_idx, row in enumerate(sheet.iter_rows(values_only=True), start=1) if row[0] is not None and "GROWTH" in row[0]]

    # loop through the cells and set the font color based on the cell value
    for row in sheet.iter_rows(min_row=2, min_col=2):
        if row[0].row not in growth_rows:
            continue
        for cell in row:
            value = cell.value
            #extracting % so we can work with the numbers
            if isinstance(value, str) and value.endswith("%"):
                try:
                    value = float(value[:-1]) / 100  # remove the % sign and divide by 100
                except ValueError:
                    continue  # skip the cell if it contains a non-numeric value
            elif not isinstance(value, (int, float)):
                continue  # skip the cell if it is not a numeric value
            if value <= 0:
                cell.font = openpyxl.styles.Font(color='FF0000')  # set font color to red for negative values
            else:
                cell.font = openpyxl.styles.Font(color='00FF00')  # set font color to green for non-negative values
    wb.save(f'financial statements ({ticker}).xlsx')
    
    #Continueing to modify next sheet
    
    company_calculation_Modification(ticker)


# In[25]:


def company_calculation_Modification(ticker):
    wb = openpyxl.load_workbook(f'financial statements ({ticker}).xlsx')
    sheet = wb["Rate Evaluation"]
    
    # loop through the cells and set the font color based on the cell value
    for row in sheet.iter_rows(min_row=2, min_col=2):
        for cell in row:
            value = cell.value
            #extracting % so we can work with the numbers
            if isinstance(value, str) and value.endswith("%"):
                try:
                    value = float(value[:-1]) / 100  # remove the % sign and divide by 100
                except ValueError:
                    continue  # skip the cell if it contains a non-numeric value
            elif not isinstance(value, (int, float)):
                continue  # skip the cell if it is not a numeric value
            if value < 0.1:
                cell.font = openpyxl.styles.Font(color='FF0000')  # set font color to red for negative values
            else:
                cell.font = openpyxl.styles.Font(color='00FF00')  # set font color to green for non-negative values
    wb.save(f'financial statements ({ticker}).xlsx')

    print('Finished Modification with :', ticker)
    print("-------------------------------------------------------------")


# In[ ]:





# In[ ]:





# In[ ]:




